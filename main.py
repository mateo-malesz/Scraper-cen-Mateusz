import requests
from bs4 import BeautifulSoup
from requests.exceptions import SSLError
import tkinter as tk
from tkinter import scrolledtext, filedialog, messagebox
import openpyxl
import re
import sys
import time
import warnings

warnings.filterwarnings("ignore", message="Unverified HTTPS request")


class DualOutput:
    def __init__(self, filename):
        self.terminal = sys.stdout
        self.log = open(filename, "w")

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        self.terminal.flush()
        self.log.flush()

    #Chyba warto dodać zamykanie pliku na końcu
    #def close(self):
        #self.log.close()

sys.stdout = DualOutput("output.txt")
new_stores = []  # Lista nowych sklepów

# Funkcja do dodawania pogrubionego tekstu
def insert_bold_text(text_widget, text):
    text_widget.insert(tk.INSERT, text, "bold")

def scrap_website(url):
    try:
        if "empis.pl" in url:
            response = requests.get(url, verify=False)
        else:
            response = requests.get(url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            website = str(url)
            website = website.strip()   #Usunięcie białych znaków z początku oraz końca linku
            website = website.replace("http://","") #Usunięcie http:// lub https:// z linku
            website = website.replace("https://","")
            website = website.split("/")    #Rozdzielenie linku na część z nazwą sklepu oraz część z nazwą produktu
            website = website[0]    #Wybranie części z nazwą sklepu
            website = website.replace("www.", "")   #usunięcie cząstki WWW gdyż potencjalnie mogłoby policzyć ten sam sklep podwójnie
            #Sprawdza, czy ten sklep jest na liście zapisanych sklepów
            if website in site:
                a = 0
                for x in site:
                    if x in website:
                        element = None  # Ustaw domyślną wartość elementu na None

                        if website == "nowaszkola.com":
                            element = soup.find_all(class_=priceid[a])
                            if element:
                                price_element = element[0]
                                return price_element.span.strip()
                        elif "id" == objtype[a]:
                            element = soup.find_all(id=priceid[a])
                        elif "class_" == objtype[a]:
                            element = soup.find_all(class_=priceid[a])
                        elif "itemprop" == objtype[a]:
                            element = soup.find_all(itemprop=priceid[a])
                        elif "data-ta" == objtype[a]:
                            element = soup.find_all('span', {'data-ta': priceid[a]})
                        elif "attr" == objtype[a]:
                            element = soup.find_all(attrs={priceid[a]: True})

                        if element and len(element) >= count[a] + 1:
                            price_element = element[count[a]]
                            if "attr" == objtype[a]:
                                return price_element.get(priceid[a], "Nie znaleziono ceny.")
                            return price_element.text.strip()
                        else:
                            return "Nie znaleziono ceny."
                    a += 1
            else: 
                #Informacja o pojawieniu się nowego sklepu
                print(f"\nZNALEZIONO NOWY SKLEP! {website}")
                #result_text.insert(tk.INSERT, f"\nZNALEZIONO NOWY SKLEP! {website}")
                if website not in new_stores:
                    new_stores.append(website)  # Dodaj nowy sklep do listy, jeśli nie jest duplikatem
        else:
            return "Nie udało się połączyć z stroną."
    except SSLError:
        print(f"Błąd SSL dla URL: {url}. Przechodzenie do następnego adresu URL.")
        return None  # Ignoruj błąd SSL i kontynuuj
    except Exception as e:
        print(f"Wystąpił błąd: {str(e)}")
        return None

def on_scrap():
    url = url_entry.get()
    url = url.strip()  # Usunięcie białych znaków z początku oraz końca linku
    result = scrap_website(url)
    if result is None:
        result = "Brak wyniku."
    result_text.insert(tk.INSERT, result + "\n")

def on_import_urls():
    filepath = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    #print(filepath)
    if not filepath:
        return

    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    result_text.delete(1.0, tk.END)
    products_checked = 0    #Licznik sprawdzonych produktów
    shops_checked = ["arante.pl"]   #Lista sprawdzonych sklepów
    cheaper, exact, expensive = 0, 0, 0   #Liczniki tańszych, cenowo równych i droższych produktów
    cheaper_offerts = 0 #Licznik tańszych ofert
    products_list.clear()
    products_list.append("ID;LINK;CENA")
    start_time = time.time()  #Zapisz czas rozpoczęcia scrapowania
    for row in sheet.iter_rows(min_row=2, values_only=True):
        ourTitle = row[0]
        product_id, our_url = row[1], row[2]
        tags = [row[3], row[4], row[5]]
        competitor_urls_list = []
        c = 6
        searched_tag = tag_entry.get()
        while c < len(row) and row[c] is not None:
            competitor_urls_list.append(str(row[c]))
            c += 1

        if product_id is None:
            print("\nKoniec pliku")
            break

        if searched_tag in tags:
            if competitor_urls_list:
                products_checked += 1
                print(f"\n--------------------------")
                print(f"\nPrzetwarzanie produktu {product_id}...")
                result = scrap_website(our_url)
                if result is not None and "Nie znaleziono ceny." not in result and "Nie udało się połączyć z stroną." not in result and "Wystąpił błąd" not in result:
                    ourPrice = result
                    ourPriceFloat = ourPrice.strip()  # Usuwamy białe znaki na początku i na końcu
                    #print(ourPriceFloat)
                    ourPriceFloat = ourPriceFloat.replace(" zł", "")  # Usuwamy niechciane znaki
                    #print(ourPriceFloat)
                    ourPriceFloat = ourPriceFloat.replace(",", ".")
                    #print(ourPriceFloat)
                    ourPriceFloat = ourPriceFloat.replace(" ", "")
                    ourPriceFloat = float(ourPriceFloat)
                    #print(ourPriceFloat)

                    print(f"ID Produktu: {product_id}, URL: {our_url}, Cena: {result}")
                    products_list.append(f"{product_id};{our_url};{result}")
                else:
                    print("Brak ceny na Arante.pl")
                    ourPrice = False

                #result_text.insert(tk.INSERT, f"ID Produktu: {product_id}, URL: {our_url}, Cena: {result}\n")

                if competitor_urls_list:
                    found = 0 # Oznacza, czy znaleziono tańszą ofertę
                    competitor_prices = []
                    for url in competitor_urls_list:
                        if url:  # Dodatkowo sprawdzamy, czy URL nie jest pusty
                            #print(f"Sprawdzam {url}")
                            result = scrap_website(url.strip())
                            concurentPrice = result
                            website = str(url)
                            website = website.strip()   #Usunięcie białych znaków z początku oraz końca linku
                            website = website.replace("http://","") #Usunięcie http:// lub https:// z linku
                            website = website.replace("https://","")
                            website = website.split("/")    #Rozdzielenie linku na część z nazwą sklepu oraz część z nazwą produktu
                            website = website[0]    #Wybranie części z nazwą sklepu
                            website = website.replace("www.", "")   #usunięcie cząstki WWW gdyż potencjalnie mogłoby policzyć ten sam sklep podwójnie
                            if website not in shops_checked:
                                shops_checked.append(website)
                            #print(concurentPrice)
                            if result is not None and "Nie znaleziono ceny." not in result and "Nie udało się połączyć z stroną." not in result and "Wystąpił błąd" not in result:
                                concurentPriceFloat = concurentPrice.strip()  #Usuwamy białe znaki na początku i na końcu
                                #print(concurentPriceFloat)
                                concurentPriceFloat = concurentPriceFloat.replace("zł", "")  #Usuwamy niechciane znaki
                                concurentPriceFloat = concurentPriceFloat.replace("PLN", "")
                                concurentPriceFloat = concurentPriceFloat.replace("Nasza cena:", "")
                                concurentPriceFloat = concurentPriceFloat.replace(" ", "")
                                concurentPriceFloat = concurentPriceFloat.replace("Cena:", "")
                                concurentPriceFloat = concurentPriceFloat.replace("Cenabrutto", "")
                                concurentPriceFloat = concurentPriceFloat.replace("zawiera23%podatku", "")
                                concurentPriceFloat = concurentPriceFloat.replace("VAT", "")
                                concurentPriceFloat = concurentPriceFloat.replace("(brutto)", "")
                                concurentPriceFloat = concurentPriceFloat.replace("brutto", "")
                                concurentPriceFloat = concurentPriceFloat.replace("z VAT", "")
                                print(concurentPriceFloat)
                                try:
                                    concurentPriceFloat = float(concurentPriceFloat)
                                except Exception as e:
                                    #print(concurentPriceFloat)
                                    concurentPriceFloat = concurentPriceFloat.replace(".", "")
                                    concurentPriceFloat = concurentPriceFloat.replace(",", ".")
                                    #print(concurentPriceFloat)
                                    concurentPriceFloat = re.sub(r'\s+', '', concurentPriceFloat)
                                    concurentPriceFloat = float(concurentPriceFloat)
                                competitor_prices.append([url.strip(), concurentPriceFloat])
                                #print(competitor_prices[0])
                                products_list.append(f"{product_id};{url.strip()};{concurentPriceFloat}")
                                try:
                                    if concurentPriceFloat < ourPriceFloat:
                                        found = 1
                                        cheaper_offerts += 1
                                except UnboundLocalError as e:
                                    pass
                                #print(f"Konkurencyjny URL dla {product_id}: {url.strip()}")
                            if result is None:
                                result = "Brak wyniku"
                    for i in range(len(competitor_prices)):
                        for a in range(len(competitor_prices) - 1):
                            if competitor_prices[a][1] > competitor_prices[a+1][1]:
                                temp = competitor_prices[a]
                                competitor_prices[a] = competitor_prices[a+1]
                                competitor_prices[a+1] = temp
                    try:
                        if competitor_prices and competitor_prices[0][1] < ourPriceFloat:
                            minPrice = competitor_prices[0][1]
                            minPriceURL = competitor_prices[0][0]
                            expensive += 1
                        elif competitor_prices and competitor_prices[0][1] == ourPriceFloat:
                            minPrice = competitor_prices[0][1]
                            minPriceURL = competitor_prices[0][0]
                            exact += 1
                        else:
                            minPrice = ourPriceFloat
                            minPriceURL = our_url
                            cheaper += 1
                        #result_text.insert(tk.INSERT, f"\n{product_id}: Nasza cena {ourPrice}, najniższa cena: {minPrice} zł w {minPriceURL}.\n")
                        # Wstaw tekst, pogrubiając {product_id}
                        insert_bold_text(result_text, f"\n{ourTitle} ({product_id})")
                        #insert_bold_text(result_text, f"\n{product_id}")
                        result_text.insert(tk.INSERT,f": \nNasza cena {ourPriceFloat:.2f} zł, najniższa cena: {minPrice:.2f} zł w {minPriceURL}.\n")
                    except UnboundLocalError as e:  #Przypadek, gdy nie ma naszej ceny (nie zaliczy produktu do statystyk w podsumowaniu)
                        if competitor_prices:   #Sprawdza, czy są ceny konkurencji
                            result_text.insert(tk.INSERT, f"\n{product_id}: Najniższa cena: {competitor_prices[0][1]:.2f} zł w {competitor_prices[0][0].strip()}")
                        else:
                            result_text.insert(tk.INSERT, f"\nBrak danych dla {product_id}.\n")
                    if found == 1 and ourPrice and competitor_prices[0][1] != 0:
                        priceDifference = ourPriceFloat - competitor_prices[0][1]
                        percentDifference = ((ourPriceFloat - competitor_prices[0][1]) / competitor_prices[0][1])*100
                        print(f"Minimalna cena dla {product_id}: {competitor_prices[0][0].strip()}, {competitor_prices[0][1]:.2f} zł, Cena w Arante {ourPriceFloat:.2f} zł wyższa o: {priceDifference:.2f} zł ({percentDifference:.2f}%)")
                    elif ourPrice:
                        print(f"\n{product_id}: CENA OK\n")
                    elif len(competitor_prices) > 0:
                        print(f"\n{product_id}: Brak ceny / nieoferowany produkt przez Arante. Najtańszy u konkurencji: {competitor_prices[0][0].strip()}, {competitor_prices[0][1]:.2f} zł\n")
                    else:
                        print(f"\n{product_id}: Brak cen / nikt nie oferuje ten produkt.\n")
                    if found == 1 or ourPrice == False:
                        print("\nReszta konkurencji:\n")
                        for c in range(1, len(competitor_prices)):
                            try:
                                if competitor_prices[c][1] < ourPriceFloat:
                                    diff = ourPriceFloat - competitor_prices[c][1]
                                    print(f"{competitor_prices[c][0].strip()}, {competitor_prices[c][1]:.2f}zł, {diff:.2f}ZŁ TANIEJ NIŻ W ARANTE")
                                else:
                                    print(f"{competitor_prices[c][0].strip()}, {competitor_prices[c][1]:.2f}zł")
                            except UnboundLocalError as e:
                                print(f"{competitor_prices[c][0].strip()}, {competitor_prices[c][1]:.2f}zł")
                    else:
                        print("\nCeny konkurencji:")
                        for c in range(len(competitor_prices)):
                            print(f"{competitor_prices[c][0].strip()}, {competitor_prices[c][1]:.2f}zł")
            else:
                print(f"Brak linków konkurencji dla produktu: {product_id}")
                #result_text.insert(tk.INSERT, f"\nBrak linków konkurencji dla produktu: {product_id}\n")
                result_text.insert(tk.INSERT, f"\nBrak linków konkurencji dla produktu: ")
                insert_bold_text(result_text, f"{ourTitle} ({product_id})\n"
                                 )
    # Tworzenie fragmentów tekstu zgodne z gramatyką języka polskiego
    if products_checked == 1:
        text_to_show = f"{products_checked} produkt w "
    elif (str(products_checked).endswith("2") or str(products_checked).endswith("3") or str(products_checked).endswith("4")) and (products_checked < 10 or products_checked > 20):
        text_to_show = f"{products_checked} produkty w "
    else:
        text_to_show = f"{products_checked} produktów w "
    if len(shops_checked) == 1:
        text_to_show2 = f"{len(shops_checked)} sklepie:\n"
    else:
        text_to_show2 = f"{len(shops_checked)} sklepach:\n"
    if cheaper == 1:
        text_to_show3 = f"TAŃSZY {cheaper} produkt"
    elif (str(cheaper).endswith("2") or str(cheaper).endswith("3") or str(cheaper).endswith("4")) and (cheaper < 10 or cheaper > 20):
        text_to_show3 = f"TAŃSZE {cheaper} produkty"
    else:
        text_to_show3 = f"TAŃSZYCH {cheaper} produktów"
    if exact == 1:
        text_to_show5 = f"W TEJ SAMEJ CENIE {exact} produkt"
    elif (str(exact).endswith("2") or str(exact).endswith("3") or str(exact).endswith("4")) and (exact < 10 or exact > 20):
        text_to_show5 = f"W TEJ SAMEJ CENIE {exact} produkty"
    else:
        text_to_show5 = f"W TEJ SAMEJ CENIE {exact} produktów"
    if expensive == 1:
        text_to_show6 = f"DROŻSZY {expensive} produkt "
    elif (str(expensive).endswith("2") or str(expensive).endswith("3") or str(expensive).endswith("4")) and (expensive < 10 or expensive > 20):
        text_to_show6 = f"DROŻSZE {expensive} produkty "
    else:
        text_to_show6 = f"DROŻSZYCH {expensive} produktów "
    if cheaper_offerts == 1:
        text_to_show4 = f"(od {cheaper_offerts} oferty konkurencji)."
    else:
        text_to_show4 = f"(od {cheaper_offerts} ofert konkurencji)."

    result_text.insert(tk.INSERT,f"\n-----------------------------")
    result_text.insert(tk.INSERT, f"\nSPRAWDZONO {text_to_show}{text_to_show2}\n{text_to_show3}\n{text_to_show5}\n{text_to_show6}{text_to_show4}\n")
    print(f"\n-----------------------------")
    print(f"\nSprawdzono {text_to_show}{text_to_show2}\n{text_to_show3}\n{text_to_show5}\n{text_to_show6}{text_to_show4}")

    if new_stores:
        print("\nLista nowych sklepów:")
        result_text.insert(tk.INSERT, f"\n-----------------------------")
        result_text.insert(tk.INSERT, "\nLista nowych sklepów:")
        for store in new_stores:
            print(store)
            result_text.insert(tk.INSERT, f"\n{store}")

    duration = time.time() - start_time  # Oblicz czas trwania
    print(f"\nCzas trwania scrapowania: {duration:.2f} sekundy")
    result_text.insert(tk.INSERT, f"\n\n-----------------------------")
    result_text.insert(tk.INSERT, f"\nCzas trwania scrapowania: {duration:.2f} sekundy")


def on_export_results():
    filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    if not filepath:
        return

    # wb = openpyxl.Workbook()
    # sheet = wb.active
    #
    # content = result_text.get("1.0", tk.END)
    # lines = content.split("\n")
    #
    # for i, line in enumerate(lines):
    #     if line:  # Pomiń puste linie
    #
    #         sheet.cell(row=i + 1, column=1, value=line)
    #
    # wb.save(filepath)
    wb = openpyxl.Workbook()
    sheet = wb.active
    i = 1
    for line in products_list:
        parts = line.split(";")
        parts[2] = parts[2].replace(",", ".")
        parts[2] = parts[2].replace("Nasza cena:", ".")
        parts[2] = parts[2].replace(" ", "")
        parts[2] = parts[2].replace("()", "")
        parts[2] = parts[2].replace("PLN", "")
        parts[2] = parts[2].replace("Cena:", "")
        parts[2] = parts[2].replace("Cenabrutto", "")
        parts[2] = parts[2].replace("zawiera23%podatku", "")
        parts[2] = parts[2].replace("VAT", "")
        parts[2] = parts[2].replace("zł", "")
        parts[2] = parts[2].replace("brutto", "")
        parts[2] = parts[2].replace("(brutto)", "")
        parts[2] = parts[2].replace("z VAT", "")
        parts[2] = re.sub(r'\s+', '', parts[2])
        if "zł" not in parts[2] and parts[2] != "CENA":
            # parts[2] = f"{parts[2]} zł"
            parts[2] = f"{parts[2]}"
        sheet.cell(row=i, column=1, value=parts[0])
        sheet.cell(row=i, column=2, value=parts[1])
        sheet.cell(row=i, column=3, value=parts[2])
        i += 1
    file = filepath.split(".")
    # file = f"{file[0]}_short.xlsx"
    file = f"{file[0]}.xlsx"
    wb.save(file)
    messagebox.showinfo("Export Complete", "Wyniki zostały zapisane do plików Excel.")

def on_show_shops():
    result_text.delete(1.0, tk.END)
    result_text.insert(tk.INSERT, "Zapisane sklepy:\n\n")
    for shop in site:
        result_text.insert(tk.INSERT, f"{shop}\n\n")

shoplist = str("Sklepy.xlsx") #Lokalizacja pliku Sklepy.xlsx
workbook = openpyxl.load_workbook(shoplist)
sh = workbook.active
i = 0
site = []       #Strona sklepu
priceid = []    #identyfikator pola z ceną
objtype = []    #typ identyfikatora(atrybut znacznika), np itemprop, id, class_
count = []      #numer znacznika(o danym identyfikatorze i typie) z ceną produktu(0 - pierwszy taki znacznik, 1 - drugi itd.)
for row in sh.iter_rows(min_row=1, values_only=True):
    site.append(str(row[0]))
    if site[i] == "":
        site.pop()
        break
    priceid.append(str(row[1]))
    objtype.append(str(row[2]))
    count.append(row[3])
    i += 1
products_list = []
# Ustawienie głównego okna aplikacji
root = tk.Tk()
root.title("Scraper Cen IXION v2024.11.14")
root.iconbitmap("ikona.ico")

tk.Label(root, text="URL:").pack(pady=5)
url_entry = tk.Entry(root, width=50)
url_entry.pack(pady=5)

scrap_button = tk.Button(root, text="Scrap Website", command=on_scrap)
scrap_button.pack(pady=5)

import_button = tk.Button(root, text="Import URLs from Excel", command=on_import_urls)
import_button.pack(pady=5)

tag_entry = tk.Entry(root, width=15)
tag_entry.pack(pady=1)

export_button = tk.Button(root, text="Export Results to Excel", command=on_export_results)
export_button.pack(pady=5)

show_button = tk.Button(root, text="Show shops", command=on_show_shops)
show_button.pack(pady=5)

result_text = scrolledtext.ScrolledText(root, height=40, width=100)
result_text.pack(pady=10)

# Dodaj styl do pogrubienia
result_text.tag_configure("bold", font=("TkDefaultFont", 10, "bold"))

#root.config(bg="#26242f")

root.mainloop()
#input("Naciśnij Enter, aby zakończyć...")

#Zamykanie pliku logów
#sys.stdout.close()